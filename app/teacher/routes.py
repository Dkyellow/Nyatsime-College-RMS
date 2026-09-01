import csv
import io
from flask import render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from datetime import datetime
from app.teacher import teacher_bp
from app.models import (
    db, User, Teacher, Student, Class, Subject, Report, Mark,
    Grade, GradeSubject, AcademicYear, AcademicTerm, TeacherSubjectClass
)
from functools import wraps
from app.services.pdf_service import invalidate_report_cache
from app.services import periods
from app.academic import calculate_grade, grade_scale_public


def teacher_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.role != 'teacher':
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


def get_assigned_subjects(teacher_id, class_id):
    """Get subjects a teacher is assigned to teach in a specific class."""
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher_id, class_id=class_id
    ).all()
    subject_ids = [a.subject_id for a in assignments]
    if not subject_ids:
        return []
    return Subject.query.filter(Subject.id.in_(subject_ids)).all()


def get_teacher_id():
    """Get the teacher record for the current user."""
    return Teacher.query.filter_by(user_id=current_user.id).first()


def get_subjects_for_grade(grade_id):
    grade = Grade.query.get(grade_id)
    return grade.subjects if grade else Subject.query.all()


def calculate_positions(class_id, term, year):
    reports = Report.query.filter_by(
        class_id=class_id, academic_term=term, academic_year=year
    ).filter(Report.status.in_(['submitted', 'approved', 'published'])).all()
    reports.sort(key=lambda r: r.average, reverse=True)
    for i, report in enumerate(reports, 1):
        report.position = i


def calculate_grade_positions(grade_id, term, year):
    grade = Grade.query.get(grade_id)
    if not grade:
        return
    class_ids = [c.id for c in grade.classes]
    if not class_ids:
        return
    reports = Report.query.filter(
        Report.class_id.in_(class_ids),
        Report.academic_term == term, Report.academic_year == year,
        Report.status.in_(['submitted', 'approved', 'published'])
    ).all()
    reports.sort(key=lambda r: r.average, reverse=True)
    for i, report in enumerate(reports, 1):
        report.grade_position = i


def invalidate_class_caches(class_id):
    ids = [r[0] for r in Report.query.with_entities(Report.id).filter_by(class_id=class_id).all()]
    for rid in ids:
        invalidate_report_cache(rid)


@teacher_bp.route('/')
@teacher_bp.route('/dashboard')
@teacher_required
def dashboard():
    teacher = Teacher.query.filter_by(user_id=current_user.id).first()
    classes = teacher.classes
    total_students = sum(len([s for s in c.students if s.is_active]) for c in classes)
    class_ids = [c.id for c in classes]
    drafts = Report.query.filter(Report.class_id.in_(class_ids), Report.status == 'draft').count() if class_ids else 0
    submitted = Report.query.filter(Report.class_id.in_(class_ids), Report.status == 'submitted').count() if class_ids else 0
    published = Report.query.filter(Report.class_id.in_(class_ids), Report.status == 'published').count() if class_ids else 0

    year_obj = periods.get_current_year()
    current_term = periods.get_current_term()
    default_term, default_year = periods.get_default_period()

    recent_results = []
    if class_ids:
        recent_results = Report.query.filter(
            Report.class_id.in_(class_ids), Report.status != 'draft'
        ).order_by(Report.updated_at.desc()).limit(8).all()

    return render_template('teacher/teacher_dashboard.html',
                           teacher=teacher, classes=classes,
                           total_students=total_students, drafts=drafts,
                           submitted=submitted, published=published,
                           current_term=current_term, on_break=current_term is None,
                           year_obj=year_obj, default_term=default_term,
                           recent_results=recent_results)


@teacher_bp.route('/class/<int:class_id>')
@teacher_required
def view_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    # Only show subjects this teacher is assigned to teach
    subjects = get_assigned_subjects(teacher.id, class_id)

    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.admission_number).all()
    grade = class_obj.grade
    years = AcademicYear.query.filter_by(is_active=True).order_by(AcademicYear.name.desc()).all()
    default_term, default_year = periods.get_default_period()
    academic_term = request.args.get('term', default_term)
    academic_year = request.args.get('year', default_year)
    reports = {r.student_id: r for r in Report.query.filter_by(
        class_id=class_id, academic_term=academic_term, academic_year=academic_year).all()}

    # Get terms for the current year for the import modal
    year_obj = AcademicYear.query.filter_by(name=academic_year).first()
    terms = sorted(year_obj.terms, key=lambda t: t.display_order) if year_obj else []

    return render_template('teacher/view_class.html',
                           class_obj=class_obj, students=students, subjects=subjects,
                           reports=reports, academic_term=academic_term,
                           academic_year=academic_year, grade=grade,
                           academic_years=years, default_term=default_term,
                           terms=terms)


@teacher_bp.route('/class/<int:class_id>/subject-marks')
@teacher_required
def subject_marks(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    # Only show assigned subjects
    subjects = get_assigned_subjects(teacher.id, class_id)
    subject_ids = [s.id for s in subjects]

    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.admission_number).all()
    years = AcademicYear.query.filter_by(is_active=True).order_by(AcademicYear.name.desc()).all()
    default_term, default_year = periods.get_default_period()
    academic_term = request.args.get('term', default_term)
    academic_year = request.args.get('year', default_year)

    subject_id = request.args.get('subject_id', type=int)
    subject = Subject.query.get(subject_id) if subject_id and subject_id in subject_ids else None

    # If no valid subject selected, redirect to first assigned subject
    if not subject and subjects:
        return redirect(url_for('teacher.subject_marks', class_id=class_id,
                                term=academic_term, year=academic_year, subject_id=subjects[0].id))

    existing = {}
    if subject:
        reports = Report.query.filter_by(class_id=class_id, academic_term=academic_term, academic_year=academic_year).all()
        for r in reports:
            mark = next((m for m in r.marks if m.subject_id == subject.id), None)
            if mark:
                existing[r.student_id] = {'score': mark.score, 'grade': mark.grade, 'status': r.status}

    return render_template('teacher/subject_marks.html',
                           class_obj=class_obj, students=students, subjects=subjects,
                           subject=subject, existing=existing,
                           academic_term=academic_term, academic_year=academic_year,
                           grading_scale=grade_scale_public(),
                           academic_years=years, default_term=default_term)


@teacher_bp.route('/class/<int:class_id>/subject-marks/save', methods=['POST'])
@teacher_required
def save_subject_marks(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    grade = class_obj.grade
    academic_term = request.form.get('term', 'Term 1')
    academic_year = request.form.get('year', str(datetime.now().year))
    subject_id = request.form.get('subject_id', type=int)

    # Verify subject is assigned to this teacher for this class
    assigned_subject_ids = [a.subject_id for a in assignments]
    if not subject_id or subject_id not in assigned_subject_ids:
        flash('You are not assigned to teach this subject in this class.', 'danger')
        return redirect(url_for('teacher.subject_marks', class_id=class_id))

    subject = Subject.query.get_or_404(subject_id)
    students = Student.query.filter_by(class_id=class_id, is_active=True).all()
    saved = 0
    for student in students:
        raw = request.form.get(f'score_{student.id}', '').strip()
        if raw == '':
            continue
        try:
            score = max(0.0, min(float(raw), float(subject.max_score or 100)))
        except ValueError:
            continue

        report = Report.query.filter_by(student_id=student.id, class_id=class_id,
                                        academic_term=academic_term, academic_year=academic_year).first()
        if not report:
            report = Report(student_id=student.id, class_id=class_id,
                            academic_term=academic_term, academic_year=academic_year, status='draft')
            db.session.add(report)
            db.session.flush()
            # Create marks only for assigned subjects
            for subj in [Subject.query.get(sid) for sid in assigned_subject_ids]:
                if subj:
                    db.session.add(Mark(report_id=report.id, subject_id=subj.id, score=0, max_score=subj.max_score))
            db.session.flush()

        if report.status in ['approved', 'published']:
            continue

        mark = Mark.query.filter_by(report_id=report.id, subject_id=subject.id).first()
        if not mark:
            mark = Mark(report_id=report.id, subject_id=subject.id, max_score=subject.max_score)
            db.session.add(mark)
        mark.score = score
        pct = (score / (mark.max_score or 100)) * 100
        mark.grade = calculate_grade(pct)
        saved += 1

        scores = [m.score for m in report.marks]
        report.total_marks = sum(scores)
        report.average = sum(scores) / len(scores) if scores else 0
        report.overall_grade = calculate_grade(report.average)

    calculate_positions(class_id, academic_term, academic_year)
    if grade:
        calculate_grade_positions(grade.id, academic_term, academic_year)

    db.session.commit()
    invalidate_class_caches(class_id)
    flash(f'{saved} mark{"s" if saved != 1 else ""} saved for {subject.name}.', 'success')
    return redirect(url_for('teacher.subject_marks', class_id=class_id,
                            term=academic_term, year=academic_year, subject_id=subject.id))


@teacher_bp.route('/report/<int:student_id>/<int:class_id>')
@teacher_required
def view_report(student_id, class_id):
    student = Student.query.get_or_404(student_id)
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    grade = class_obj.grade
    default_term, default_year = periods.get_default_period()
    academic_term = request.args.get('term', default_term)
    academic_year = request.args.get('year', default_year)

    report = Report.query.filter_by(student_id=student_id, class_id=class_id,
                                    academic_term=academic_term, academic_year=academic_year).first()
    if not report:
        report = Report(student_id=student_id, class_id=class_id,
                        academic_term=academic_term, academic_year=academic_year, status='draft')
        db.session.add(report)
        db.session.flush()
        # Only create marks for assigned subjects
        assigned_subject_ids = [a.subject_id for a in assignments]
        for subj_id in assigned_subject_ids:
            subj = Subject.query.get(subj_id)
            if subj:
                db.session.add(Mark(report_id=report.id, subject_id=subj.id, score=0, max_score=subj.max_score))
        db.session.commit()

    # Only show assigned subjects
    subjects = get_assigned_subjects(teacher.id, class_id)
    years = AcademicYear.query.filter_by(is_active=True).order_by(AcademicYear.name.desc()).all()

    return render_template('teacher/view_report.html',
                           student=student, class_obj=class_obj, report=report,
                           subjects=subjects, grading_scale=grade_scale_public(),
                           grade=grade, academic_years=years,
                           default_term=default_term)


@teacher_bp.route('/report/update/<int:report_id>', methods=['POST'])
@teacher_required
def update_report(report_id):
    report = Report.query.get_or_404(report_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=report.class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    # Only allow editing marks for assigned subjects
    assigned_subject_ids = [a.subject_id for a in assignments]

    if report.status in ['approved', 'published']:
        flash('Cannot edit a report that has been approved or published.', 'danger')
        return redirect(url_for('teacher.view_report',
                                student_id=report.student_id, class_id=report.class_id,
                                term=report.academic_term, year=report.academic_year))

    for mark in report.marks:
        # Only update marks for assigned subjects
        if mark.subject_id not in assigned_subject_ids:
            continue
        score_key = f'score_{mark.subject_id}'
        score = request.form.get(score_key, 0, type=float)
        max_score = mark.max_score or 100
        score = max(0, min(score, max_score))
        mark.score = score
        mark.grade = calculate_grade((score / max_score) * 100 if max_score else 0)

    scores = [m.score for m in report.marks]
    total = sum(scores)
    report.total_marks = total
    report.average = total / len(scores) if scores else 0
    report.overall_grade = calculate_grade(report.average)
    report.teacher_comment = request.form.get('teacher_comment', '')

    db.session.commit()
    invalidate_report_cache(report.id)
    flash('Marks saved successfully!', 'success')
    return redirect(url_for('teacher.view_report',
                            student_id=report.student_id, class_id=report.class_id,
                            term=report.academic_term, year=report.academic_year))


@teacher_bp.route('/report/submit/<int:report_id>', methods=['POST'])
@teacher_required
def submit_report(report_id):
    report = Report.query.get_or_404(report_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=report.class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    if report.status in ['approved', 'published']:
        flash('Report is already approved or published.', 'danger')
        return redirect(url_for('teacher.view_report',
                                student_id=report.student_id, class_id=report.class_id,
                                term=report.academic_term, year=report.academic_year))

    # Only allow submitting marks for assigned subjects
    assigned_subject_ids = [a.subject_id for a in assignments]

    # Save marks from form before submitting
    for mark in report.marks:
        # Only update marks for assigned subjects
        if mark.subject_id not in assigned_subject_ids:
            continue
        score_key = f'score_{mark.subject_id}'
        raw = request.form.get(score_key, '')
        if raw == '':
            continue
        try:
            score = float(raw)
        except (ValueError, TypeError):
            continue
        max_score = mark.max_score or 100
        score = max(0, min(score, max_score))
        mark.score = score
        mark.grade = calculate_grade((score / max_score) * 100 if max_score else 0)

    scores = [m.score for m in report.marks]
    total = sum(scores)
    report.total_marks = total
    report.average = total / len(scores) if scores else 0
    report.overall_grade = calculate_grade(report.average)
    report.teacher_comment = request.form.get('teacher_comment', '') or report.teacher_comment

    report.status = 'submitted'
    report.submitted_at = datetime.utcnow()

    calculate_positions(report.class_id, report.academic_term, report.academic_year)
    if report.class_obj and report.class_obj.grade:
        calculate_grade_positions(report.class_obj.grade.id, report.academic_term, report.academic_year)

    db.session.commit()
    invalidate_report_cache(report.id)
    flash('Report submitted for approval!', 'success')
    return redirect(url_for('teacher.view_class', class_id=report.class_id))


@teacher_bp.route('/upload-marks/<int:class_id>', methods=['POST'])
@teacher_required
def upload_marks(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    academic_term = request.form.get('term', 'Term 1')
    academic_year = request.form.get('year', str(datetime.now().year))
    grade = class_obj.grade

    # Only process assigned subjects
    assigned_subject_ids = [a.subject_id for a in assignments]
    assigned_subjects = Subject.query.filter(Subject.id.in_(assigned_subject_ids)).all()

    if 'file' not in request.files or request.files['file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('teacher.view_class', class_id=class_id, term=academic_term, year=academic_year))

    file = request.files['file']
    try:
        if file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        else:
            flash('Unsupported file format. Please use CSV or Excel.', 'danger')
            return redirect(url_for('teacher.view_class', class_id=class_id, term=academic_term, year=academic_year))

        # Build subject map only for assigned subjects
        subject_map = {s.code.upper(): s.id for s in assigned_subjects}
        subject_map.update({s.name.upper(): s.id for s in assigned_subjects})
        subject_map.update({f"{s.code.upper()} - {s.name.upper()}": s.id for s in assigned_subjects})

        count = 0
        for row in reader:
            admission_number = str(row.get('admission_number', '')).strip()
            student = Student.query.filter_by(admission_number=admission_number, class_id=class_id).first()
            if not student:
                continue

            report = Report.query.filter_by(student_id=student.id, class_id=class_id,
                                            academic_term=academic_term, academic_year=academic_year).first()
            if not report:
                report = Report(student_id=student.id, class_id=class_id,
                                academic_term=academic_term, academic_year=academic_year, status='draft')
                db.session.add(report)
                db.session.flush()
                # Create marks only for assigned subjects
                for subject in assigned_subjects:
                    db.session.add(Mark(report_id=report.id, subject_id=subject.id, score=0, max_score=subject.max_score))
                db.session.flush()

            if report.status in ['approved', 'published']:
                continue

            for key, value in row.items():
                if key and key.strip().upper() in ['ADMISSION_NUMBER', 'STUDENT', 'FIRST_NAME', 'LAST_NAME']:
                    continue
                if key and value is not None:
                    subject_id = subject_map.get(key.strip().upper())
                    if subject_id:
                        subj = Subject.query.get(subject_id)
                        if not subj:
                            continue
                        try:
                            score = float(value)
                            max_score_val = subj.max_score or 100
                            score = max(0, min(score, max_score_val))
                            mark = Mark.query.filter_by(report_id=report.id, subject_id=subject_id).first()
                            if mark:
                                mark.score = score
                                pct = (score / max_score_val) * 100
                                mark.grade = calculate_grade(pct)
                                count += 1
                        except (ValueError, TypeError):
                            continue

            total = sum(m.score for m in report.marks)
            report.total_marks = total
            report.average = total / len(report.marks) if report.marks else 0
            report.overall_grade = calculate_grade(report.average)

        calculate_positions(class_id, academic_term, academic_year)
        if grade:
            calculate_grade_positions(grade.id, academic_term, academic_year)
        db.session.commit()
        flash(f'Marks uploaded successfully! {count} marks updated.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error uploading marks: {str(e)}', 'danger')

    return redirect(url_for('teacher.view_class', class_id=class_id, term=academic_term, year=academic_year))


@teacher_bp.route('/download-template/<int:class_id>')
@teacher_required
def download_template(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    # Only include assigned subjects
    subjects = get_assigned_subjects(teacher.id, class_id)
    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.admission_number).all()

    output = io.StringIO()
    writer = csv.writer(output)
    headers = ['admission_number', 'first_name', 'last_name'] + [f"{s.code} - {s.name}" for s in subjects]
    writer.writerow(headers)
    for student in students:
        writer.writerow([student.admission_number, student.first_name, student.last_name] + [''] * len(subjects))

    response = Response(output.getvalue())
    filename = f"marks_template_{class_obj.name.replace(' ', '_')}.csv"
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@teacher_bp.route('/export-marks/<int:class_id>')
@teacher_required
def export_marks(class_id):
    class_obj = Class.query.get_or_404(class_id)
    teacher = get_teacher_id()

    # Verify teacher is assigned to this class
    assignments = TeacherSubjectClass.query.filter_by(
        teacher_id=teacher.id, class_id=class_id
    ).all()
    if not assignments:
        flash('You are not assigned to teach in this class.', 'danger')
        return redirect(url_for('teacher.dashboard'))

    academic_term = request.args.get('term', 'Term 1')
    academic_year = request.args.get('year', str(datetime.now().year))

    # Only export assigned subjects
    subjects = get_assigned_subjects(teacher.id, class_id)
    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.admission_number).all()

    output = io.StringIO()
    writer = csv.writer(output)
    headers = ['admission_number', 'first_name', 'last_name']
    for subject in subjects:
        headers.append(subject.code)
    writer.writerow(headers)

    for student in students:
        report = Report.query.filter_by(student_id=student.id, class_id=class_id,
                                        academic_term=academic_term, academic_year=academic_year).first()
        row = [student.admission_number, student.first_name, student.last_name]
        for subject in subjects:
            score = ''
            if report:
                mark = Mark.query.filter_by(report_id=report.id, subject_id=subject.id).first()
                if mark:
                    score = mark.score
            row.append(score)
        writer.writerow(row)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=marks_{class_obj.name}_{academic_term}_{academic_year}.csv'}
    )
