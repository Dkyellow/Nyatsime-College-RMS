import csv
import os
import uuid
import io
from flask import render_template, redirect, url_for, flash, request, Response, jsonify, current_app
from flask_login import login_required, current_user
from datetime import datetime
from app.admin import admin_bp
from app.models import (
    db, User, Admin, Teacher, Student, Class, Subject, Report, Mark,
    Grade, GradeSubject, AcademicYear,
    AcademicTerm, SchoolSetting, TeacherSubjectClass, AuditLog,
    StudentSubject, StudentPromotion
)
from functools import wraps
from app.services.pdf_service import invalidate_report_cache
from app.services import periods
from app.academic import calculate_grade, generate_username, slugify_name, FIXED_FORMS


def log_action(action, resource_type=None, resource_id=None, details=None):
    """Log a user action to the audit trail."""
    from flask import request
    entry = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        details=details,
        ip_address=request.remote_addr
    )
    db.session.add(entry)


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== DASHBOARD ====================

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    total_students = Student.query.filter_by(is_active=True).count()
    total_teachers = Teacher.query.count()
    total_classes = Class.query.count()
    total_subjects = Subject.query.count()
    pending_reports = Report.query.filter_by(status='submitted').count()
    generated_reports = Report.query.filter(Report.status.in_(['approved', 'published'])).count()

    recent_results = Report.query.filter(
        Report.marks.any(), Report.status != 'draft'
    ).order_by(Report.updated_at.desc()).limit(6).all()

    recent_activity = Report.query.order_by(Report.created_at.desc()).limit(6).all()

    attention_reports = Report.query.filter(
        Report.status.in_(['submitted', 'approved', 'published'])
    ).order_by(Report.average.asc()).limit(40).all()

    attention, seen = [], set()
    for r in attention_reports:
        if r.student_id not in seen:
            seen.add(r.student_id)
            attention.append((r.student, r.average))
        if len(attention) >= 5:
            break

    current_year_obj = periods.get_current_year()
    current_term = periods.get_current_term(current_year_obj)
    on_break = current_term is None

    default_term, default_year = periods.get_default_period()

    return render_template('dashboard.html',
                           total_students=total_students,
                           total_teachers=total_teachers,
                           total_classes=total_classes,
                           total_subjects=total_subjects,
                           pending_reports=pending_reports,
                           generated_reports=generated_reports,
                           recent_results=recent_results,
                           recent_activity=recent_activity,
                           attention=attention,
                           current_year=current_year_obj.name,
                           current_term=current_term,
                           on_break=on_break,
                           default_term=default_term,
                           quick_stats={
                               'draft': Report.query.filter_by(status='draft').count(),
                               'approved': Report.query.filter_by(status='approved').count(),
                           })


# ==================== ACADEMIC CALENDAR ====================

@admin_bp.route('/academic-calendar')
@admin_required
def academic_calendar():
    years = periods.available_years()
    selected = request.args.get('year', '', type=str)
    year_name = selected if selected in years else (str(datetime.now().year) if str(datetime.now().year) in years else (years[0] if years else str(datetime.now().year)))

    year = periods.ensure_year(year_name)
    db.session.commit()
    terms = sorted(year.terms, key=lambda t: t.display_order)
    today = datetime.now().date()
    current_term = periods.get_current_term()

    return render_template('academic_calendar.html',
                           years=years,
                           year=year,
                           terms=terms,
                           selected_year=year_name,
                           today=today,
                           current_term=current_term)


@admin_bp.route('/academic-calendar/save/<int:year_id>', methods=['POST'])
@admin_required
def save_academic_calendar(year_id):
    year = AcademicYear.query.get_or_404(year_id)
    for term in year.terms:
        prefix = f'term_{term.id}'
        start_raw = request.form.get(f'{prefix}_start', '').strip()
        end_raw = request.form.get(f'{prefix}_end', '').strip()
        try:
            term.start_date = datetime.strptime(start_raw, '%Y-%m-%d').date() if start_raw else None
            term.end_date = datetime.strptime(end_raw, '%Y-%m-%d').date() if end_raw else None
        except ValueError:
            flash(f'Invalid date supplied for {term.name}. Use the date picker.', 'danger')
            return redirect(url_for('admin.academic_calendar', year=year.name))

        if term.start_date and term.end_date and term.end_date < term.start_date:
            flash(f'{term.name}: end date cannot be before the start date.', 'danger')
            db.session.rollback()
            return redirect(url_for('admin.academic_calendar', year=year.name))

    db.session.commit()
    invalidate_all_caches()
    flash(f'Academic calendar for {year.name} saved.', 'success')
    return redirect(url_for('admin.academic_calendar', year=year.name))


@admin_bp.route('/academic-calendar/prepare-year', methods=['POST'])
@admin_required
def prepare_year():
    name = request.form.get('year_name', '').strip()
    normalized = name[:4] if len(name) >= 4 else ''
    if not normalized.isdigit():
        flash('Enter a valid single year, e.g. 2027.', 'danger')
        return redirect(url_for('admin.academic_calendar'))

    existing = AcademicYear.query.filter_by(name=normalized).first()
    if existing:
        flash(f'Academic year {normalized} already exists.', 'info')
        return redirect(url_for('admin.academic_calendar', year=normalized))

    periods.ensure_year(normalized)
    db.session.commit()
    flash(f'Academic year {normalized} prepared with default term dates - adjust them below.', 'success')
    return redirect(url_for('admin.academic_calendar', year=normalized))


# ==================== API ENDPOINTS ====================

@admin_bp.route('/api/classes-by-grade/<int:grade_id>')
@admin_required
def api_classes_by_grade(grade_id):
    classes = Class.query.filter_by(grade_id=grade_id).all()
    return jsonify([{'id': c.id, 'name': c.name} for c in classes])


@admin_bp.route('/api/subjects-by-grade/<int:grade_id>')
@admin_required
def api_subjects_by_grade(grade_id):
    grade = Grade.query.get_or_404(grade_id)
    subjects = grade.subjects
    return jsonify([{'id': s.id, 'name': s.name, 'code': s.code, 'max_score': s.max_score} for s in subjects])


@admin_bp.route('/api/student-username-preview')
@admin_required
def api_username_preview():
    first = request.args.get('first_name', '')
    last = request.args.get('last_name', '')

    def exists(username):
        return User.query.filter_by(username=username).first() is not None

    base = f"{slugify_name(first)}{slugify_name(last)}"
    preview = generate_username(first, last, exists)
    return jsonify({'base': base, 'username': preview})


# ==================== FORM SUBJECTS (curriculum) ====================
# Forms themselves are fixed reference data - only subject assignment is managed.

@admin_bp.route('/grade-subjects')
@admin_required
def grade_subjects():
    from app.models import Grade as G
    periods.ensure_fixed_grades(G)
    grade_id = request.args.get('grade_id', type=int)
    grades_list = G.query.order_by(G.display_order).all()

    all_subjects = Subject.query.order_by(Subject.name).all()
    selected_grade = None
    assigned_subject_ids = []
    if not grade_id and grades_list:
        grade_id = grades_list[0].id
    if grade_id:
        selected_grade = db.session.get(G, grade_id)
        assigned_subject_ids = [s.id for s in selected_grade.subjects]

    return render_template('grade_subjects.html',
                           grades=grades_list,
                           all_subjects=all_subjects,
                           selected_grade=selected_grade,
                           assigned_subject_ids=assigned_subject_ids)


@admin_bp.route('/grade-subjects/update/<int:grade_id>', methods=['POST'])
@admin_required
def update_grade_subjects(grade_id):
    grade = Grade.query.get_or_404(grade_id)
    subject_ids = request.form.getlist('subject_ids')

    grade.grade_subjects.clear()
    for sid in subject_ids:
        gs = GradeSubject(grade_id=grade.id, subject_id=int(sid))
        db.session.add(gs)

    db.session.commit()
    flash('Subjects for this form updated successfully!', 'success')
    return redirect(url_for('admin.grade_subjects', grade_id=grade_id))


# ==================== REPORT TEMPLATES ====================

@admin_bp.route('/report-templates')
@admin_required
def report_templates():
    templates = ReportTemplate.query.all()
    return render_template('report_templates.html', report_templates=templates)


@admin_bp.route('/report-templates/add', methods=['POST'])
@admin_required
def add_report_template():
    template = ReportTemplate(
        name=request.form.get('name'),
        template_type='secondary',
        description=request.form.get('description')
    )
    db.session.add(template)
    db.session.commit()
    flash('Report template added successfully!', 'success')
    return redirect(url_for('admin.report_templates'))


@admin_bp.route('/report-templates/edit/<int:id>', methods=['POST'])
@admin_required
def edit_report_template(id):
    template = db.session.get(ReportTemplate, id) or abort_404(ReportTemplate, id)
    template.name = request.form.get('name')
    template.description = request.form.get('description')
    db.session.commit()
    flash('Report template updated successfully!', 'success')
    return redirect(url_for('admin.report_templates'))


@admin_bp.route('/report-templates/delete/<int:id>', methods=['POST'])
@admin_required
def delete_report_template(id):
    template = db.session.get(ReportTemplate, id)
    if template:
        db.session.delete(template)
        db.session.commit()
    flash('Report template deleted successfully!', 'success')
    return redirect(url_for('admin.report_templates'))


def abort_404(model, id):
    from flask import abort
    obj = db.session.get(model, id)
    if obj is None:
        abort(404)
    return obj


# ==================== STUDENT MANAGEMENT ====================

@admin_bp.route('/students')
@admin_required
def students():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    class_id = request.args.get('class_id', type=int)

    query = Student.query.filter_by(is_active=True)
    if search:
        query = query.filter(
            db.or_(
                Student.first_name.ilike(f'%{search}%'),
                Student.last_name.ilike(f'%{search}%'),
                Student.admission_number.ilike(f'%{search}%')
            )
        )
    if class_id:
        query = query.filter(Student.class_id == class_id)

    students_list = query.order_by(Student.admission_number).paginate(page=page, per_page=10)
    classes = Class.query.order_by(Class.name).all()
    return render_template('students.html', students=students_list, classes=classes,
                           search=search, selected_class=class_id)


@admin_bp.route('/students/<int:id>')
@admin_required
def student_profile(id):
    student = Student.query.get_or_404(id)
    reports = Report.query.filter_by(student_id=id).order_by(
        Report.academic_year.desc(), Report.academic_term).all()
    return render_template('student_profile.html', student=student, reports=reports)


def _create_student_user(first_name, last_name, password=None):
    """Create a login account for a student with a generated unique username."""
    def exists(username):
        return User.query.filter_by(username=username).first() is not None

    username = generate_username(first_name, last_name, exists)
    user = User(username=username, email=None, role='student')
    user.set_password(password or 'student123')
    db.session.add(user)
    db.session.flush()
    return user


@admin_bp.route('/students/add', methods=['POST'])
@admin_required
def add_student():
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    admission_number = request.form.get('admission_number', '').strip()

    if not first_name or not last_name or not admission_number:
        flash('First name, surname and admission number are required.', 'danger')
        return redirect(url_for('admin.students'))
    if Student.query.filter_by(admission_number=admission_number).first():
        flash('A student with that admission number already exists.', 'danger')

        return redirect(url_for('admin.students'))

    date_of_birth = request.form.get('date_of_birth')
    gender = request.form.get('gender')
    class_id = request.form.get('class_id')
    password = request.form.get('password', '').strip()

    user = _create_student_user(first_name, last_name, password or None)

    student = Student(
        user_id=user.id,
        first_name=first_name,
        last_name=last_name,
        admission_number=admission_number,
        date_of_birth=datetime.strptime(date_of_birth, '%Y-%m-%d').date() if date_of_birth else None,
        gender=gender,
        class_id=int(class_id) if class_id else None,
    )
    db.session.add(student)
    log_action('create', 'student', None, f'Added student {first_name} {last_name} ({admission_number})')
    db.session.commit()
    flash(f"Student enrolled! Portal login: {user.username}", 'success')
    return redirect(url_for('admin.students'))


@admin_bp.route('/students/edit/<int:id>', methods=['POST'])
@admin_required
def edit_student(id):
    student = Student.query.get_or_404(id)
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    student.first_name = first_name
    student.last_name = last_name
    student.admission_number = request.form.get('admission_number', '').strip()
    student.gender = request.form.get('gender')
    student.class_id = int(request.form.get('class_id')) if request.form.get('class_id') else None
    dob = request.form.get('date_of_birth')
    if dob:
        student.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date()

    password = request.form.get('password', '').strip()
    if student.user_id:
        user = db.session.get(User, student.user_id)
        if password:
            user.set_password(password)
    elif first_name and last_name:
        user = _create_student_user(first_name, last_name, password or None)
        student.user_id = user.id

    log_action('update', 'student', student.id, f'Updated student {student.first_name} {student.last_name}')
    db.session.commit()
    flash('Student updated successfully!', 'success')
    return redirect(request.referrer or url_for('admin.students'))


@admin_bp.route('/students/delete/<int:id>', methods=['POST'])
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    student.is_active = False
    if student.user_id:
        user = db.session.get(User, student.user_id)
        if user:
            user.is_active = False
    log_action('delete', 'student', student.id, f'Deactivated student {student.first_name} {student.last_name}')
    db.session.commit()
    flash('Student removed from the roll (portal access disabled).', 'success')
    return redirect(url_for('admin.students'))


@admin_bp.route('/students/upload', methods=['POST'])
@admin_required
def upload_students():
    if 'file' not in request.files or request.files['file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('admin.students'))

    file = request.files['file']
    try:
        if file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        else:
            flash('Unsupported file format. Please use CSV or Excel.', 'danger')
            return redirect(url_for('admin.students'))

        count, accounts = 0, 0
        for row in reader:
            first_name = (row.get('first_name') or '').strip()
            last_name = (row.get('last_name') or '').strip()
            admission_number = (row.get('admission_number') or '').strip()
            gender = (row.get('gender') or '').strip()
            class_name = (row.get('class') or '').strip()
            dob = (row.get('date_of_birth') or '').strip()

            if not first_name or not last_name or not admission_number:
                continue
            if Student.query.filter_by(admission_number=admission_number).first():
                continue

            class_obj = Class.query.filter_by(name=class_name).first() if class_name else None

            user = _create_student_user(first_name, last_name)
            accounts += 1

            student = Student(
                user_id=user.id,
                first_name=first_name,
                last_name=last_name,
                admission_number=admission_number,
                gender=gender if gender in ['Male', 'Female'] else None,
                class_id=class_obj.id if class_obj else None,
                date_of_birth=datetime.strptime(dob, '%Y-%m-%d').date() if dob else None
            )
            db.session.add(student)
            count += 1

        db.session.commit()
        msg = f'{count} students imported'
        if accounts:
            msg += f' ({accounts} portal logins created - initial password: student123)'
        flash(msg + '.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing students: {str(e)}', 'danger')

    return redirect(url_for('admin.students'))


@admin_bp.route('/students/export')
@admin_required
def export_students():
    students = Student.query.filter_by(is_active=True).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['first_name', 'last_name', 'admission_number', 'gender', 'form', 'class', 'portal_username', 'date_of_birth'])

    for student in students:
        writer.writerow([
            student.first_name,
            student.last_name,
            student.admission_number,
            student.gender or '',
            student.class_obj.grade.name if student.class_obj and student.class_obj.grade else '',
            student.class_obj.name if student.class_obj else '',
            student.user.username if student.user else '',
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else ''
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=students_export.csv'}
    )


@admin_bp.route('/reports/export')
@admin_required
def export_reports():
    class_id = request.args.get('class_id', type=int)
    term = request.args.get('term', 'Term 1')
    year = request.args.get('year', str(datetime.now().year))

    query = Report.query.filter_by(academic_term=term, academic_year=year)
    if class_id:
        query = query.filter_by(class_id=class_id)
    reports = query.all()

    subjects = Subject.query.all()

    output = io.StringIO()
    writer = csv.writer(output)
    headers = ['admission_number', 'student_name', 'class', 'term', 'year', 'total', 'average', 'grade', 'position', 'status']
    for subject in subjects:
        headers.append(subject.code)
    writer.writerow(headers)

    for report in reports:
        student = report.student
        row = [
            student.admission_number,
            f'{student.first_name} {student.last_name}',
            report.class_obj.name if report.class_obj else '',
            report.academic_term,
            report.academic_year,
            report.total_marks,
            round(report.average, 1),
            report.overall_grade,
            report.position or '',
            report.status
        ]
        for subject in subjects:
            mark = Mark.query.filter_by(report_id=report.id, subject_id=subject.id).first()
            row.append(mark.score if mark else '')
        writer.writerow(row)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=reports_{term}_{year}.csv'}
    )


# ==================== TEACHER MANAGEMENT ====================

@admin_bp.route('/teachers')
@admin_required
def teachers():
    page = request.args.get('page', 1, type=int)
    teachers = Teacher.query.paginate(page=page, per_page=10)
    return render_template('teachers.html', teachers=teachers)


@admin_bp.route('/teachers/add', methods=['POST'])
@admin_required
def add_teacher():
    username = request.form.get('username', '').strip()
    email = request.form.get('email', '').strip()
    if User.query.filter_by(username=username).first():
        flash('Username already taken.', 'danger')
        return redirect(url_for('admin.teachers'))
    if email and User.query.filter_by(email=email).first():
        flash('Email already in use.', 'danger')
        return redirect(url_for('admin.teachers'))

    user = User(username=username, email=email or None, role='teacher')
    user.set_password(request.form.get('password'))
    db.session.add(user)
    db.session.flush()

    teacher = Teacher(
        user_id=user.id,
        first_name=request.form.get('first_name'),
        last_name=request.form.get('last_name'),
        phone=request.form.get('phone'),
        employee_id=request.form.get('employee_id')
    )
    db.session.add(teacher)
    log_action('create', 'teacher', None, f'Added teacher {teacher.first_name} {teacher.last_name}')
    db.session.commit()
    flash('Teacher added successfully!', 'success')
    return redirect(url_for('admin.teachers'))


@admin_bp.route('/teachers/edit/<int:id>', methods=['POST'])
@admin_required
def edit_teacher(id):
    teacher = Teacher.query.get_or_404(id)
    new_username = request.form.get('username')
    if new_username != teacher.user.username:
        existing = User.query.filter_by(username=new_username).first()
        if existing:
            flash('Username already taken.', 'danger')
            return redirect(url_for('admin.teachers'))
        teacher.user.username = new_username
    teacher.first_name = request.form.get('first_name')
    teacher.last_name = request.form.get('last_name')
    teacher.phone = request.form.get('phone')
    teacher.employee_id = request.form.get('employee_id')
    password = request.form.get('password')
    if password:
        teacher.user.set_password(password)
    log_action('update', 'teacher', teacher.id, f'Updated teacher {teacher.first_name} {teacher.last_name}')
    db.session.commit()
    flash('Teacher updated successfully!', 'success')
    return redirect(url_for('admin.teachers'))


@admin_bp.route('/teachers/delete/<int:id>', methods=['POST'])
@admin_required
def delete_teacher(id):
    teacher = Teacher.query.get_or_404(id)
    user = teacher.user
    log_action('delete', 'teacher', teacher.id, f'Deleted teacher {teacher.first_name} {teacher.last_name}')
    db.session.delete(teacher)
    db.session.delete(user)
    db.session.commit()
    flash('Teacher removed successfully!', 'success')
    return redirect(url_for('admin.teachers'))


# ==================== USERS & ROLES ====================

@admin_bp.route('/users')
@admin_required
def users():
    role = request.args.get('role', '')
    query = User.query
    if role in ('admin', 'teacher', 'student'):
        query = query.filter_by(role=role)
    users_list = query.order_by(User.role, User.username).all()
    return render_template('users.html', users=users_list, selected_role=role)


@admin_bp.route('/users/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_user(id):
    user = db.session.get(User, id)
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.users'))
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'warning')
        return redirect(url_for('admin.users'))
    user.is_active = not user.is_active
    db.session.commit()
    flash(f"Account '{user.username}' {'activated' if user.is_active else 'deactivated'}.", 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/reset-password/<int:id>', methods=['POST'])
@admin_required
def reset_user_password(id):
    user = db.session.get(User, id)
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('admin.users'))
    password = request.form.get('password')
    if password and len(password) >= 6:
        user.set_password(password)
        db.session.commit()
        flash(f"Password reset for '{user.username}'.", 'success')
    else:
        flash('Password must be at least 6 characters.', 'danger')
    return redirect(url_for('admin.users'))


# ==================== CLASS MANAGEMENT ====================

@admin_bp.route('/classes')
@admin_required
def classes():
    periods.ensure_fixed_grades(Grade)
    grade_id = request.args.get('grade_id', type=int)
    grades_list = Grade.query.order_by(Grade.display_order).all()

    query = Class.query
    if grade_id:
        query = query.filter_by(grade_id=grade_id)
    classes_list = query.order_by(Class.name).all()

    teachers = Teacher.query.all()
    return render_template('classes.html', classes=classes_list, teachers=teachers,
                           grades=grades_list, selected_grade=grade_id)


@admin_bp.route('/classes/add', methods=['POST'])
@admin_required
def add_class():
    name = request.form.get('name')
    section = request.form.get('section')
    grade_id = request.form.get('grade_id', type=int)

    class_obj = Class(name=name, section=section, grade_id=grade_id)
    db.session.add(class_obj)
    db.session.flush()

    # Handle teacher-subject assignments
    teacher_ids = request.form.getlist('teacher_ids')
    subject_ids = request.form.getlist('subject_ids')
    for teacher_id in teacher_ids:
        for subject_id in subject_ids:
            assignment = TeacherSubjectClass(
                teacher_id=int(teacher_id),
                class_id=class_obj.id,
                subject_id=int(subject_id)
            )
            db.session.add(assignment)

    log_action('create', 'class', class_obj.id, f'Created class {name}')
    db.session.commit()
    flash('Class/stream created successfully!', 'success')
    return redirect(url_for('admin.classes'))


@admin_bp.route('/classes/edit/<int:id>', methods=['POST'])
@admin_required
def edit_class(id):
    class_obj = Class.query.get_or_404(id)
    class_obj.name = request.form.get('name')
    class_obj.section = request.form.get('section')
    class_obj.grade_id = request.form.get('grade_id', type=int)

    # Clear existing assignments and recreate
    TeacherSubjectClass.query.filter_by(class_id=class_obj.id).delete()
    teacher_ids = request.form.getlist('teacher_ids')
    subject_ids = request.form.getlist('subject_ids')
    for teacher_id in teacher_ids:
        for subject_id in subject_ids:
            assignment = TeacherSubjectClass(
                teacher_id=int(teacher_id),
                class_id=class_obj.id,
                subject_id=int(subject_id)
            )
            db.session.add(assignment)

    log_action('update', 'class', class_obj.id, f'Updated class {class_obj.name}')
    db.session.commit()
    flash('Class/stream updated successfully!', 'success')
    return redirect(url_for('admin.classes'))


@admin_bp.route('/classes/delete/<int:id>', methods=['POST'])
@admin_required
def delete_class(id):
    class_obj = Class.query.get_or_404(id)
    if class_obj.students:
        flash('Cannot delete a class that still has students assigned.', 'danger')
        return redirect(url_for('admin.classes'))
    db.session.delete(class_obj)
    db.session.commit()
    flash('Class/stream deleted successfully!', 'success')
    return redirect(url_for('admin.classes'))


# ==================== SUBJECT MANAGEMENT ====================

@admin_bp.route('/subjects')
@admin_required
def subjects():
    subjects = Subject.query.order_by(Subject.name).all()
    grades_list = Grade.query.order_by(Grade.display_order).all()
    return render_template('subjects.html', subjects=subjects, grades=grades_list)


@admin_bp.route('/subjects/add', methods=['POST'])
@admin_required
def add_subject():
    code = request.form.get('code', '').upper()
    if Subject.query.filter_by(code=code).first():
        flash('A subject with that code already exists.', 'danger')
        return redirect(url_for('admin.subjects'))
    subject = Subject(
        name=request.form.get('name'),
        code=code,
        max_score=request.form.get('max_score', 100, type=int)
    )
    db.session.add(subject)
    db.session.commit()
    flash('Subject added successfully!', 'success')
    return redirect(url_for('admin.subjects'))


@admin_bp.route('/subjects/edit/<int:id>', methods=['POST'])
@admin_required
def edit_subject(id):
    subject = Subject.query.get_or_404(id)
    subject.name = request.form.get('name')
    subject.code = request.form.get('code', '').upper()
    subject.max_score = request.form.get('max_score', 100, type=int)
    db.session.commit()
    flash('Subject updated successfully!', 'success')
    return redirect(url_for('admin.subjects'))


@admin_bp.route('/subjects/delete/<int:id>', methods=['POST'])
@admin_required
def delete_subject(id):
    subject = Subject.query.get_or_404(id)
    db.session.delete(subject)
    db.session.commit()
    flash('Subject deleted successfully!', 'success')
    return redirect(url_for('admin.subjects'))


# ==================== REPORT MANAGEMENT ====================

@admin_bp.route('/reports')
@admin_required
def reports():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    class_id = request.args.get('class_id', type=int)
    year = request.args.get('year', '')

    query = Report.query
    if status:
        query = query.filter_by(status=status)
    if class_id:
        query = query.filter(Report.class_id == class_id)
    if year:
        query = query.filter(Report.academic_year == year)

    reports_list = query.order_by(Report.updated_at.desc()).paginate(page=page, per_page=12)
    classes = Class.query.order_by(Class.name).all()
    years = sorted({y[0] for y in Report.query.with_entities(Report.academic_year).distinct()}, reverse=True)
    default_term, default_year = periods.get_default_period()

    return render_template('reports.html', reports=reports_list, current_status=status,
                           classes=classes, years=years,
                           selected_class=class_id, selected_year=year,
                           default_term=default_term, default_year=default_year)


@admin_bp.route('/reports/approve/<int:id>', methods=['POST'])
@admin_required
def approve_report(id):
    report = Report.query.get_or_404(id)
    report.status = 'approved'
    report.approved_at = datetime.utcnow()
    admin_comment = request.form.get('admin_comment')
    if admin_comment:
        report.admin_comment = admin_comment
    log_action('approve', 'report', report.id, f'Approved report for student #{report.student_id}')
    db.session.commit()
    invalidate_report_cache(report.id)
    flash('Report approved successfully!', 'success')
    return redirect(request.referrer or url_for('admin.reports'))


@admin_bp.route('/reports/publish/<int:id>', methods=['POST'])
@admin_required
def publish_report(id):
    report = Report.query.get_or_404(id)
    report.status = 'published'
    report.published_at = datetime.utcnow()
    log_action('publish', 'report', report.id, f'Published report for student #{report.student_id}')
    db.session.commit()
    invalidate_report_cache(report.id)
    flash('Report published - now visible to the student portal.', 'success')
    return redirect(request.referrer or url_for('admin.reports'))


# ==================== SCHOOL SETTINGS ====================

ALLOWED_LOGO_EXTENSIONS = {'png', 'jpg', 'jpeg', 'svg', 'webp'}


def _allowed_logo(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_LOGO_EXTENSIONS


@admin_bp.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    # All text-based setting keys managed by this page
    text_keys = [
        'school_name', 'school_short_name', 'school_motto',
        'school_address', 'school_city', 'school_country',
        'school_phone', 'school_email', 'school_website',
        'primary_color', 'accent_color', 'report_footer',
    ]

    if request.method == 'POST':
        # Save all text fields
        for key in text_keys:
            SchoolSetting.set(key, request.form.get(key, '').strip())

        # Handle logo removal
        if request.form.get('remove_logo'):
            old_filename = SchoolSetting.get('logo_filename', '')
            if old_filename:
                old_path = os.path.join(
                    current_app.root_path, 'static', 'uploads', old_filename
                )
                try:
                    os.remove(old_path)
                except OSError:
                    pass
            SchoolSetting.set('logo_filename', '')

        # Handle logo upload
        logo_file = request.files.get('logo_file')
        if logo_file and logo_file.filename and _allowed_logo(logo_file.filename):
            # Remove previous logo if one exists
            old_filename = SchoolSetting.get('logo_filename', '')
            if old_filename:
                old_path = os.path.join(
                    current_app.root_path, 'static', 'uploads', old_filename
                )
                try:
                    os.remove(old_path)
                except OSError:
                    pass

            ext = logo_file.filename.rsplit('.', 1)[1].lower()
            new_filename = f'school_logo_{uuid.uuid4().hex[:8]}.{ext}'
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            logo_file.save(os.path.join(upload_dir, new_filename))
            SchoolSetting.set('logo_filename', new_filename)
        elif logo_file and logo_file.filename and not _allowed_logo(logo_file.filename):
            flash('Unsupported logo format. Please use PNG, JPG, SVG or WebP.', 'danger')

        db.session.commit()
        invalidate_all_caches()
        flash('School settings saved successfully.', 'success')
        return redirect(url_for('admin.settings'))

    all_keys = text_keys + ['logo_filename']
    values = {key: SchoolSetting.get(key) for key in all_keys}
    return render_template('settings.html', values=values)


def invalidate_all_caches():
    ids = [r[0] for r in Report.query.with_entities(Report.id).all()]
    for rid in ids:
        invalidate_report_cache(rid)


# ==================== AUDIT LOGS ====================

@admin_bp.route('/audit-logs')
@admin_required
def audit_logs():
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action', '')
    user_id_filter = request.args.get('user_id', type=int)

    query = AuditLog.query
    if action_filter:
        query = query.filter_by(action=action_filter)
    if user_id_filter:
        query = query.filter_by(user_id=user_id_filter)

    logs = query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=20)
    users = User.query.order_by(User.username).all()

    actions = db.session.query(AuditLog.action).distinct().all()
    actions = [a[0] for a in actions]

    return render_template('admin/audit_logs.html', logs=logs, users=users,
                           actions=actions, selected_action=action_filter,
                           selected_user=user_id_filter)


# ==================== STUDENT PERFORMANCE AUDIT ====================

@admin_bp.route('/performance-audit')
@admin_required
def performance_audit():
    year = request.args.get('year', str(datetime.now().year))
    term = request.args.get('term', 'Term 1')
    grade_id = request.args.get('grade_id', type=int)
    class_id = request.args.get('class_id', type=int)

    grades_list = Grade.query.order_by(Grade.display_order).all()
    classes_list = Class.query.order_by(Class.name).all()

    # Get all reports for the selected period
    report_query = Report.query.filter_by(academic_year=year, academic_term=term)
    if grade_id:
        class_ids = [c.id for c in Grade.query.get(grade_id).classes]
        report_query = report_query.filter(Report.class_id.in_(class_ids))
    if class_id:
        report_query = report_query.filter_by(class_id=class_id)

    reports = report_query.filter(Report.status.in_(['submitted', 'approved', 'published'])).all()

    # Class averages
    class_stats = {}
    for report in reports:
        cid = report.class_id
        if cid not in class_stats:
            class_stats[cid] = {'sum': 0, 'count': 0, 'class': report.class_obj}
        class_stats[cid]['sum'] += report.average
        class_stats[cid]['count'] += 1

    class_averages = []
    for cid, data in class_stats.items():
        avg = data['sum'] / data['count'] if data['count'] else 0
        class_averages.append({
            'class': data['class'],
            'average': round(avg, 1),
            'student_count': data['count']
        })
    class_averages.sort(key=lambda x: x['average'], reverse=True)

    # Top students per class
    top_students = {}
    for report in reports:
        cid = report.class_id
        if cid not in top_students:
            top_students[cid] = []
        top_students[cid].append(report)
    for cid in top_students:
        top_students[cid].sort(key=lambda r: r.average, reverse=True)
        top_students[cid] = top_students[cid][:5]

    # Subject performance
    subject_stats = {}
    for report in reports:
        for mark in report.marks:
            sid = mark.subject_id
            if sid not in subject_stats:
                subject_stats[sid] = {'sum': 0, 'count': 0, 'max': 0, 'min': 100, 'subject': mark.subject}
            if mark.max_score:
                pct = (mark.score / mark.max_score) * 100
            else:
                pct = mark.score
            subject_stats[sid]['sum'] += pct
            subject_stats[sid]['count'] += 1
            subject_stats[sid]['max'] = max(subject_stats[sid]['max'], pct)
            subject_stats[sid]['min'] = min(subject_stats[sid]['min'], pct)

    subject_averages = []
    for sid, data in subject_stats.items():
        avg = data['sum'] / data['count'] if data['count'] else 0
        subject_averages.append({
            'subject': data['subject'],
            'average': round(avg, 1),
            'highest': round(data['max'], 1),
            'lowest': round(data['min'], 1),
            'count': data['count']
        })
    subject_averages.sort(key=lambda x: x['average'], reverse=True)

    # Overall stats
    all_averages = [r.average for r in reports]
    overall_avg = sum(all_averages) / len(all_averages) if all_averages else 0
    overall_highest = max(all_averages) if all_averages else 0
    overall_lowest = min(all_averages) if all_averages else 0

    return render_template('admin/performance_audit.html',
                           class_averages=class_averages,
                           top_students=top_students,
                           subject_averages=subject_averages,
                           overall_avg=round(overall_avg, 1),
                           overall_highest=round(overall_highest, 1),
                           overall_lowest=round(overall_lowest, 1),
                           total_reports=len(reports),
                           selected_year=year, selected_term=term,
                           selected_grade=grade_id, selected_class=class_id,
                           grades=grades_list, classes=classes_list)


# ==================== TEACHER-SUBJECT-CLASS ASSIGNMENTS ====================

@admin_bp.route('/teacher-assignments')
@admin_required
def teacher_assignments():
    teachers = Teacher.query.order_by(Teacher.last_name).all()
    classes = Class.query.order_by(Class.name).all()
    subjects = Subject.query.order_by(Subject.name).all()

    assignments = TeacherSubjectClass.query.all()
    assignment_map = {}
    for a in assignments:
        key = (a.teacher_id, a.class_id)
        if key not in assignment_map:
            assignment_map[key] = []
        assignment_map[key].append(a.subject)

    return render_template('admin/teacher_assignments.html',
                           teachers=teachers, classes=classes, subjects=subjects,
                           assignment_map=assignment_map)


@admin_bp.route('/teacher-assignments/save', methods=['POST'])
@admin_required
def save_teacher_assignments():
    teacher_id = request.form.get('teacher_id', type=int)
    class_id = request.form.get('class_id', type=int)
    subject_ids = request.form.getlist('subject_ids')

    if not teacher_id or not class_id:
        flash('Please select a teacher and a class.', 'danger')
        return redirect(url_for('admin.teacher_assignments'))

    # Clear existing assignments for this teacher-class combo
    TeacherSubjectClass.query.filter_by(teacher_id=teacher_id, class_id=class_id).delete()

    # Add new assignments
    for subject_id in subject_ids:
        assignment = TeacherSubjectClass(
            teacher_id=teacher_id,
            class_id=class_id,
            subject_id=int(subject_id)
        )
        db.session.add(assignment)

    teacher = Teacher.query.get(teacher_id)
    class_obj = Class.query.get(class_id)
    log_action('update', 'teacher_assignment', teacher_id,
               f'Updated subject assignments for {teacher.first_name} {teacher.last_name} in {class_obj.name}')
    db.session.commit()
    flash('Teacher-subject assignments saved successfully!', 'success')
    return redirect(url_for('admin.teacher_assignments'))


# ==================== CLASS LIST IMPORT ====================

@admin_bp.route('/import-class-list', methods=['GET', 'POST'])
@admin_required
def import_class_list():
    if request.method == 'GET':
        classes = Class.query.order_by(Class.name).all()
        return render_template('admin/import_class_list.html', classes=classes)

    class_id = request.form.get('class_id', type=int)
    academic_year = request.form.get('academic_year', str(datetime.now().year))

    if not class_id:
        flash('Please select a class.', 'danger')
        return redirect(url_for('admin.import_class_list'))

    class_obj = Class.query.get_or_404(class_id)

    if 'file' not in request.files or request.files['file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('admin.import_class_list'))

    file = request.files['file']
    try:
        if file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        else:
            flash('Unsupported file format. Please use CSV or Excel.', 'danger')
            return redirect(url_for('admin.import_class_list'))

        count = 0
        accounts_created = 0
        skipped = 0

        for row in reader:
            first_name = (row.get('first_name') or '').strip()
            last_name = (row.get('last_name') or '').strip()
            admission_number = (row.get('admission_number') or '').strip()
            gender = (row.get('gender') or '').strip()
            dob = (row.get('date_of_birth') or '').strip()

            if not first_name or not last_name or not admission_number:
                continue

            # Check if student already exists
            existing = Student.query.filter_by(admission_number=admission_number).first()
            if existing:
                skipped += 1
                continue

            # Create user account
            def exists(username):
                return User.query.filter_by(username=username).first() is not None

            username = generate_username(first_name, last_name, exists)
            password = f"{slugify_name(first_name)}123"
            user = User(username=username, email=None, role='student')
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            accounts_created += 1

            # Create student
            student = Student(
                user_id=user.id,
                first_name=first_name,
                last_name=last_name,
                admission_number=admission_number,
                gender=gender if gender in ['Male', 'Female'] else None,
                date_of_birth=datetime.strptime(dob, '%Y-%m-%d').date() if dob else None,
                class_id=class_id,
                is_active=True
            )
            db.session.add(student)
            count += 1

        db.session.commit()
        log_action('import', 'student', class_id,
                   f'Imported {count} students to {class_obj.name} ({accounts_created} accounts created)')

        msg = f'{count} students imported to {class_obj.name}.'
        if accounts_created:
            msg += f' {accounts_created} portal accounts created.'
        if skipped:
            msg += f' {skipped} duplicate admission numbers skipped.'
        flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error importing students: {str(e)}', 'danger')

    return redirect(url_for('admin.students'))


@admin_bp.route('/download-class-list-template')
@admin_required
def download_class_list_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['first_name', 'last_name', 'admission_number', 'gender', 'date_of_birth'])
    writer.writerow(['John', 'Doe', 'ADM001', 'Male', '2010-01-15'])
    writer.writerow(['Jane', 'Smith', 'ADM002', 'Female', '2010-03-22'])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=class_list_template.csv'}
    )


# ==================== STUDENT PROMOTION ====================

@admin_bp.route('/promote-students', methods=['GET', 'POST'])
@admin_required
def promote_students():
    if request.method == 'GET':
        classes = Class.query.order_by(Class.name).all()
        grades = Grade.query.order_by(Grade.display_order).all()

        # Get current year students by grade
        grade_students = {}
        for grade in grades:
            grade_students[grade.id] = {
                'grade': grade,
                'students': [],
                'count': 0
            }
            for cls in grade.classes:
                students = Student.query.filter_by(class_id=cls.id, is_active=True).all()
                grade_students[grade.id]['students'].extend(students)
                grade_students[grade.id]['count'] += len(students)

        # Define promotion paths
        promotion_paths = {
            'Form 1': 'Form 2',
            'Form 2': 'Form 3',
            'Form 3': 'Form 4',
            'Form 4': None,  # Graduates
            'Lower 6': 'Upper 6',
            'Upper 6': None,  # Graduates
        }

        return render_template('admin/promote_students.html',
                               classes=classes, grades=grades,
                               grade_students=grade_students,
                               promotion_paths=promotion_paths)

    # POST - Process promotion
    promotion_type = request.form.get('promotion_type', 'auto')
    academic_year = request.form.get('academic_year', str(datetime.now().year))
    target_year = str(int(academic_year) + 1)

    # Ensure target year exists
    target_year_obj = AcademicYear.query.filter_by(name=target_year).first()
    if not target_year_obj:
        target_year_obj = AcademicYear(name=target_year, is_active=True)
        db.session.add(target_year_obj)
        db.session.flush()
        # Create terms for new year
        from app.services import periods
        term_dates = periods.default_term_dates(target_year)
        for i, term_name in enumerate(['Term 1', 'Term 2', 'Term 3'], start=1):
            sd, ed = term_dates[term_name]
            db.session.add(AcademicTerm(
                academic_year_id=target_year_obj.id, name=term_name,
                display_order=i, start_date=sd, end_date=ed,
            ))

    promotion_paths = {
        'Form 1': 'Form 2',
        'Form 2': 'Form 3',
        'Form 3': 'Form 4',
        'Form 4': None,
        'Lower 6': 'Upper 6',
        'Upper 6': None,
    }

    promoted = 0
    graduated = 0
    needs_manual = []

    # Get selected student IDs
    student_ids = request.form.getlist('student_ids')

    if promotion_type == 'auto':
        # Auto-promote all eligible students
        for grade in Grade.query.all():
            next_form = promotion_paths.get(grade.name)
            if next_form is None:
                # Graduates - skip
                for cls in grade.classes:
                    students = Student.query.filter_by(class_id=cls.id, is_active=True).all()
                    for student in students:
                        graduated += 1
                continue

            # Find or create target grade
            target_grade = Grade.query.filter_by(name=next_form).first()
            if not target_grade:
                continue

            for cls in grade.classes:
                students = Student.query.filter_by(class_id=cls.id, is_active=True).all()
                for student in students:
                    # Find or create target class in target grade
                    target_class = Class.query.filter_by(
                        grade_id=target_grade.id, name=cls.name
                    ).first()
                    if not target_class:
                        # Create new class with same name in target grade
                        target_class = Class(
                            name=cls.name,
                            section=cls.section,
                            grade_id=target_grade.id
                        )
                        db.session.add(target_class)
                        db.session.flush()

                    # Record promotion
                    promotion = StudentPromotion(
                        student_id=student.id,
                        from_class_id=cls.id,
                        to_class_id=target_class.id,
                        from_grade_id=grade.id,
                        to_grade_id=target_grade.id,
                        academic_year=academic_year,
                        promotion_type='auto'
                    )
                    db.session.add(promotion)

                    # Update student
                    student.class_id = target_class.id
                    promoted += 1

    elif promotion_type == 'manual' and student_ids:
        # Promote selected students
        for sid in student_ids:
            student = Student.query.get(int(sid))
            if not student or not student.class_obj:
                continue

            current_grade = student.class_obj.grade
            next_form = promotion_paths.get(current_grade.name)

            if next_form is None:
                graduated += 1
                continue

            target_grade = Grade.query.filter_by(name=next_form).first()
            if not target_grade:
                continue

            # Find or create target class
            target_class = Class.query.filter_by(
                grade_id=target_grade.id, name=student.class_obj.name
            ).first()
            if not target_class:
                target_class = Class(
                    name=student.class_obj.name,
                    section=student.class_obj.section,
                    grade_id=target_grade.id
                )
                db.session.add(target_class)
                db.session.flush()

            promotion = StudentPromotion(
                student_id=student.id,
                from_class_id=student.class_id,
                to_class_id=target_class.id,
                from_grade_id=current_grade.id,
                to_grade_id=target_grade.id,
                academic_year=academic_year,
                promotion_type='manual'
            )
            db.session.add(promotion)
            student.class_id = target_class.id
            promoted += 1

    log_action('promote', 'student', None,
               f'Promoted {promoted} students, {graduated} graduated')

    db.session.commit()
    flash(f'Promotion complete: {promoted} students promoted, {graduated} students graduated.', 'success')
    return redirect(url_for('admin.promote_students'))


# ==================== FORM 3 TRANSITION ====================

@admin_bp.route('/form3-transition', methods=['GET', 'POST'])
@admin_required
def form3_transition():
    if request.method == 'GET':
        # Get Form 2 classes
        form2_grade = Grade.query.filter_by(name='Form 2').first()
        form2_classes = Class.query.filter_by(grade_id=form2_grade.id).all() if form2_grade else []

        # Get Form 3 classes (target)
        form3_grade = Grade.query.filter_by(name='Form 3').first()
        form3_classes = Class.query.filter_by(grade_id=form3_grade.id).all() if form3_grade else []

        # Get Form 2 students
        form2_students = []
        for cls in form2_classes:
            students = Student.query.filter_by(class_id=cls.id, is_active=True).all()
            for s in students:
                form2_students.append({'student': s, 'class': cls})

        # Available streams for Form 3
        streams = [
            ('Science', 'Form 3 Science'),
            ('Arts', 'Form 3 Arts'),
            ('Commercials', 'Form 3 Commerce'),
        ]

        # Get subjects for each stream
        form3_grade = Grade.query.filter_by(name='Form 3').first()
        form3_subjects = form3_grade.subjects if form3_grade else []

        return render_template('admin/form3_transition.html',
                               form2_students=form2_students,
                               form3_classes=form3_classes,
                               streams=streams,
                               form3_subjects=form3_subjects)

    # POST - Process transition
    academic_year = request.form.get('academic_year', str(datetime.now().year))
    form3_grade = Grade.query.filter_by(name='Form 3').first()

    if not form3_grade:
        flash('Form 3 grade not found.', 'danger')
        return redirect(url_for('admin.form3_transition'))

    transitions = request.form.getlist('transitions')
    promoted = 0

    for transition in transitions:
        if not transition:
            continue
        # Format: student_id:stream
        parts = transition.split(':')
        if len(parts) != 2:
            continue
        student_id = int(parts[0])
        stream = parts[1]

        student = Student.query.get(student_id)
        if not student:
            continue

        # Find or create target class
        class_name = f'Form 3 {stream}'
        target_class = Class.query.filter_by(
            grade_id=form3_grade.id, name=class_name
        ).first()
        if not target_class:
            target_class = Class(
                name=class_name,
                section=stream,
                grade_id=form3_grade.id
            )
            db.session.add(target_class)
            db.session.flush()

        # Record promotion
        promotion = StudentPromotion(
            student_id=student.id,
            from_class_id=student.class_id,
            to_class_id=target_class.id,
            from_grade_id=Grade.query.filter_by(name='Form 2').first().id,
            to_grade_id=form3_grade.id,
            academic_year=academic_year,
            promotion_type='manual',
            notes=f'Stream: {stream}'
        )
        db.session.add(promotion)

        # Update student class
        student.class_id = target_class.id
        promoted += 1

    log_action('promote', 'student', None,
               f'Form 3 transition: {promoted} students placed into streams')

    db.session.commit()
    flash(f'Form 3 transition complete: {promoted} students placed into streams.', 'success')
    return redirect(url_for('admin.form3_transition'))
